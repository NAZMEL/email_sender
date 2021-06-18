from openpyxl import load_workbook

# class for download emails from excel file
class Excel_work:
    def __init__(self, file):
        self.emails = []
        self.wb = load_workbook(file)
        self.sheet = self.wb['Receivers']

    def get_emails(self):
        for i in range(1, len(self.sheet['A'])):
            self.emails.append(self.sheet.cell(row=i, column = 1).value)

        return self.emails
     